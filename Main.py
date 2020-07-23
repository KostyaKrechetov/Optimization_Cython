import os
import time
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from datetime import datetime, timedelta, date
from math import sqrt, isnan
from scipy.optimize import differential_evolution
import gpse
import lhsmdu
import openpyxl
import random
import re

from Calculations.Welltest import Pwf_Ql_Qt

Q_liq_ = datetimes_Q_liq_ = date_diffs = deltas = difference = percentage = cumul_summ_diff = cumul_summ_diff_rel = None


def write_excel(ind):
    global Q_liq_, datetimes_Q_liq_, date_diffs, deltas, difference, percentage, cumul_summ_diff

    for col in range(len(deltas)):
        cell_abs = sheet_abs.cell(row=ind, column=int(deltas[col]) + 1)
        cell_abs.value = round(difference[col], 2)
        cell_rel = sheet_rel.cell(row=ind, column=int(deltas[col]) + 1)
        cell_rel.value = round(percentage[col], 2)
        cell_cumul = sheet_cumul.cell(row=ind, column=int(deltas[col]) + 1)
        cell_cumul.value = round(cumul_summ_diff[col], 2)
        cell_cumul_rel = sheet_cumul_rel.cell(row=ind, column=int(deltas[col]) + 1)
        cell_cumul_rel.value = round(cumul_summ_diff_rel[col], 2)
        cell_Q_liq_act = sheet_Q_liq_act.cell(row=ind, column=int(deltas[col]) + 1)
        cell_Q_liq_act.value = round(Q_liq_[col], 2)
        cell_Q_liq_mod = sheet_Q_liq_mod.cell(row=ind, column=int(deltas[col]) + 1)
        cell_Q_liq_mod.value = round(Q_comparison[col], 2)


def save_cumul(*args):
    global Q_liq_, datetimes_Q_liq_, date_diffs, deltas, difference, cumul_summ_diff, cumul_summ_diff_rel
    if len(Q_comparison) < 1:
        return 0
    cumul_summ_real = []
    cumul_summ_model = []
    cumul_summ_diff = []
    cumul_summ_diff_rel = []
    for i in range(1, len(datetimes_Q_liq_) + 1):
        cumul_summ_real.append(np.sum(Q_liq_[:i]))
        cumul_summ_model.append(np.sum(Q_comparison[:i]))
        cumul_summ_diff.append(abs(np.sum(Q_comparison[:i] - Q_liq_[:i])))
        cumul_summ_diff_rel.append(cumul_summ_diff[i - 1] / cumul_summ_real[i - 1] * 100)

    fig, ax = plt.subplots()
    ax.plot(deltas, cumul_summ_real, label='Реальные данные')
    ax.plot(deltas, cumul_summ_model, label='Модель')
    ax.set_ylabel('Накопленная добыча, м3')
    ax.set_xlabel('Сутки')
    ax2 = ax.twinx()
    ax2.plot(deltas, cumul_summ_diff, label='Накопленная разность, м3', linestyle="--", color='royalblue')
    ax2.set_ylabel('Накопленная разность, м3')
    fig.autofmt_xdate()
    ax.legend(loc=2)
    ax2.legend(loc=4)

    plt.savefig(
        os.path.join(result_directory, str(args[0]) + 'накопленная добыча' + '.png'),
        dpi=300)
    plt.close(fig)


def save_hist(*args):
    global Q_liq_, datetimes_Q_liq_, date_diffs, deltas, difference, percentage
    if len(Q_comparison) < 1:
        return 0
    Q_liq_ = Q_liq[end_Q_liq:]
    datetimes_Q_liq_ = datetimes_Q_liq[end_Q_liq:]
    date_diffs = [elem - (datetimes_P_bh[-1] - timedelta(days=comparison_period)) for elem in datetimes_Q_liq_]
    deltas = [elem.total_seconds() / 3600 / 24 for elem in date_diffs]
    difference = np.abs(np.subtract(Q_liq_, Q_comparison[:]))
    percentage = np.zeros(len(difference))
    for ii in range(0, len(difference)):
        percentage[ii] = (1 - min(Q_liq_[ii], Q_comparison[ii]) / max(Q_liq_[ii], Q_comparison[ii])) * 100
    ind = np.arange(1, len(deltas) + 1)  # the x locations for the groups
    width = 0.35  # the width of the bars

    more10 = 0
    for ii in range(0, len(percentage)):
        if percentage[ii] > 10:
            more10 = more10 + 1
    more10per = more10 / len(percentage) * 100

    fig, ax = plt.subplots()
    line1 = ax.bar(np.array(deltas) - width / 2, difference, width,
                   label='Абсолютная ошибка')
    ax2 = ax.twinx()
    line2 = ax2.bar(np.array(deltas) + width / 2, percentage, width,
                    label='Относительная ошибка', color='lightsalmon')

    # Add some text for labels, title and custom x-axis tick labels, etc.
    ax.set_xlabel('> 10%: ' + str(round(more10per, 2)) + '% из всех точек')
    ax.set_ylabel('Абсолютная, м3/сут')
    ax.set_ylim(0, max(difference) * 1.2)
    ax2.set_ylim(0, max(percentage) * 1.2)
    ax2.set_ylabel('Относительная, %')
    ax.set_title('Ошибка адаптации')
    ax.set_xticks(ind[::2])
    # ax.set_xticks(np.arange(1, 91))
    # ax.set_xticklabels(('G1', 'G2', 'G3', 'G4', 'G5'))
    fig.tight_layout()
    ax.legend((line1, line2), ('Абсолютная ошибка', 'Относительная ошибка'), loc=0)

    plt.savefig(os.path.join(result_directory, str(args[0]) + 'histogram' + '.png'),
                dpi=300)
    plt.close(fig)


def save_plot(*args, prediction=False):
    fig, axes = plt.subplots(ncols=1, nrows=2, sharex=True)
    color = 'red'
    axes[0].set_xlabel('time')
    axes[0].set_ylabel('Q_жид, м3/сутки', color=color)
    axes[0].set_zorder(1)
    axes[0].tick_params(axis='y', labelcolor=color)
    axes[0].tick_params(axis='x', labelbottom=True)
    axes[0].grid(b=True, which='both', axis='both')

    axes[1].set_ylabel('P_bh, атм', color='green')
    axes[1].tick_params(axis='y', labelcolor='green')
    axes[1].grid(b=True, which='both', axis='both')

    color = 'red'
    line1 = axes[0].plot(datetimes_Q_liq[start_Q_liq:], Q_liq[start_Q_liq:], marker='.', color=color, label='Q_liq')
    if prediction:
        Q_ = Q.tolist()
        Q_c = Q_comparison.tolist()
        a = Q_.copy()
        # a.append(Q_c[0])
        line2 = axes[0].plot(datetimes_Q_liq[start_Q_liq:start_Q_liq + len(Q_)], a, '.', linestyle="--",
                             color='darkred', label='Адаптация Q_liq')
        line3 = axes[0].plot(datetimes_Q_liq[start_Q_liq + len(Q_):],  Q_c, linestyle="--",
                             color='royalblue', label='Прогноз Q_liq')
    else:
        line2 = axes[0].plot(datetimes_Q_liq[start_Q_liq:end_Q_liq], Q, '.', linestyle="--", color='darkred',
                             label='Адаптация Q_liq')

    # Настройка легенды
    if prediction:
        lns = line1 + line2 + line3
    else:
        lns = line1 + line2
    labels = [l.get_label() for l in lns]
    axes[0].legend(lns, labels, loc=0)

    axes[1].plot(datetimes_P_bh[start_P_bh:], P_bh[start_P_bh:], marker='.', color='green', label='P_bh')
    axes[1].legend()
    fig.autofmt_xdate()
    # plt.show()
    plt.savefig(os.path.join(result_directory, str(args[0]) + '.png'), dpi=300)
    plt.close(fig)


def loss_func(x, *args):
    global time_points, input_list, Q, weight, x_, best, num_calls

    delta_dates_ = delta_dates + comparison_delta_dates
    pw = bh_pressures
    pw.append(pw[-1])
    time_moments = [np.sum(delta_dates_[:i + 1]) for i in range(len(delta_dates_))]
    time_moments.insert(0, 0)
    delta_p = [bh_pressures[0] - bh_pressures[i] for i in range(1, len(bh_pressures))]
    delta_p.append(delta_p[-1])
    slope = [-(pw[i + 1] - pw[i]) / delta_dates_[i] for i in range(len(delta_dates_))]
    free_term = [x[4] - (pw[i] * time_moments[i + 1] - pw[i + 1] * time_moments[i]) / delta_dates_[i] for i in
                 range(len(delta_dates_))]

    input_list = [(delta_dates_[i] / 24, delta_p[i], slope[i], free_term[i]) for i in range(len(delta_dates_))]

    # input_list = [((delta_dates + comparison_delta_dates)[i] / 24, x[4] - bh_pressures[i]) for i in
    #               range(len(delta_dates + comparison_delta_dates))]
    num_calls += 1
    Q = np.array([])
    for t in time_points:
        # q = Pwf_Ql_Qt(t=t, Regimes=input_list, k=x[0], regimes_flag=1,
        #                       skin=x[1], h=5.96, ct= 0.000047936014476, mu=0.338932567825029, bl=1.02787536,
        #                       rw=0.108, phi=0.15, Wellmodel=2,
        #                       ReservoirShape=0, BoundaryType=0, xf=94.28,
        #                       kv_kh=x[2], Fc=10000,
        #                       w=x[3], l=x[4], wf=0.5, lf=0.5)
        # q = Pwf_Ql_Qt(t=t, Regimes=input_list, k=x[0], regimes_flag=1,
        #               skin=x[1], h=11.35, ct=0.0000557258307023158, mu=0.356131019863818, bl=1.04370981614035,
        #               rw=0.108, phi=0.145, Wellmodel=WM,
        #               ReservoirShape=0, BoundaryType=0, xf=94.28,
        #               kv_kh=x[2], Fc=10000,
        #               w=x[3], l=x[4], wf=0.5, lf=0.5)
        # q = Pwf_Ql_Qt(t=t, Regimes=input_list, k=x[0], regimes_flag=1,
        #               skin=x[1], h=10, ct=5e-5, mu=1.5, bl=1.2,
        #               rw=0.1, phi=0.2, Wellmodel=WM,
        #               ReservoirShape=0, BoundaryType=0, xf=94.28,
        #               kv_kh=x[2], Fc=10000,
        #               w=x[3], l=x[4], wf=0.5, lf=0.5)
        try:
            q = Pwf_Ql_Qt(t=t, Regimes=input_list, k=x[0], regimes_flag=2,
                          skin=x[1], h=res_prm[7], ct=res_prm[8], mu=res_prm[5], bl=res_prm[6],
                          rw=res_prm[3], phi=res_prm[4], Wellmodel=WM[res_prm[1]],
                          ReservoirShape=0, BoundaryType=BoundType, xf=x[5],
                          Fc=x[6],
                          w=x[2], l=x[3], wf=0.5, lf=0.5)
        except:
            q = Pwf_Ql_Qt(t=t, Regimes=input_list, k=x[0], regimes_flag=2,
                          skin=x[1], h=res_prm[7], ct=res_prm[8], mu=res_prm[5], bl=res_prm[6],
                          rw=res_prm[3], phi=res_prm[4], Wellmodel=WM[res_prm[1]],
                          ReservoirShape=0, BoundaryType=BoundType,
                          w=x[2], l=x[3], wf=0.5, lf=0.5)
        Q = np.append(Q, q)
    dev = (Q - Q_liq[start_Q_liq:end_Q_liq]) * (Q - Q_liq[start_Q_liq:end_Q_liq])
    loss = np.sqrt(np.sum((dev * weight)) / considered_length)
    print('loss: ' + str(loss) + '  num: ' + str(num_calls))
    # losses.append(loss)
    if loss <= best:
        best = loss
        x_ = x
    return loss


def check_convergence():
    def check_conv(x0, convergence):
        val = loss_func(x0)
    return check_conv


def loss_func_maximize(x, *args):
    temp = loss_func(x)
    return -temp


def loss_func_maximize_P_init_only(p_init_, *args):
    global time_points, input_list, Q, weight, x_, best, num_calls, x

    delta_dates_ = delta_dates + comparison_delta_dates
    pw = bh_pressures
    pw.append(pw[-1])
    time_moments = [np.sum(delta_dates_[:i + 1]) for i in range(len(delta_dates_))]
    time_moments.insert(0, 0)
    delta_p = [bh_pressures[0] - bh_pressures[i] for i in range(1, len(bh_pressures))]
    delta_p.append(delta_p[-1])
    slope = [-(pw[i + 1] - pw[i]) / delta_dates_[i] for i in range(len(delta_dates_))]
    free_term = [p_init_[0] - (pw[i] * time_moments[i + 1] - pw[i + 1] * time_moments[i]) / delta_dates_[i] for i in
                 range(len(delta_dates_))]

    input_list = [(delta_dates_[i] / 24, delta_p[i], slope[i], free_term[i]) for i in range(len(delta_dates_))]

    num_calls += 1
    Q = np.array([])
    for t in time_points:
        try:
            q = Pwf_Ql_Qt(t=t, Regimes=input_list, k=x[0], regimes_flag=2,
                          skin=x[1], h=res_prm[7], ct=res_prm[8], mu=res_prm[5], bl=res_prm[6],
                          rw=res_prm[3], phi=res_prm[4], Wellmodel=WM[res_prm[1]],
                          ReservoirShape=0, BoundaryType=BoundType, xf=x[5],
                          Fc=x[6],
                          w=x[2], l=x[3], wf=0.5, lf=0.5)
        except:
            q = Pwf_Ql_Qt(t=t, Regimes=input_list, k=x[0], regimes_flag=2,
                          skin=x[1], h=res_prm[7], ct=res_prm[8], mu=res_prm[5], bl=res_prm[6],
                          rw=res_prm[3], phi=res_prm[4], Wellmodel=WM[res_prm[1]],
                          ReservoirShape=0, BoundaryType=BoundType,
                          w=x[2], l=x[3], wf=0.5, lf=0.5)
        Q = np.append(Q, q)
    dev = (Q - Q_liq[start_Q_liq:end_Q_liq]) * (Q - Q_liq[start_Q_liq:end_Q_liq])
    loss = np.sqrt(np.sum((dev * weight)) / considered_length)
    print('p_init:  loss: ' + str(loss) + '  num: ' + str(num_calls))
    return -loss


xl_tech_regime = pd.ExcelFile(
    'Проверка на реальных данных\\Отдельное (месторожд)\\Техрежим, ОАО  Газпромнефть-Ноябрьскнефтегаз , май 2020.xlsx')
df_tech = xl_tech_regime.parse('Режим')
df_tech = df_tech[9:]
well_nums = df_tech.iloc[:, [4]].values.tolist()
well_nums = [elem[0] for elem in well_nums]
well_nums = [elem for elem in well_nums if elem == elem]    # избавление от nan


xl_pechat = pd.ExcelFile(
        'Проверка на реальных данных\\Отдельное (месторожд)\\Печать справочника физико-химических свойств пласта.xlsx')
df_pechat = xl_pechat.parse()
df_pechat = df_pechat[12:]

res_prms = []
res_prm_line = []
for ind, w_num in enumerate(well_nums):
    w_type = df_tech.iat[ind, 5]
    plast_name = df_tech.iat[ind, 7]
    res_prm_line.extend([w_num, w_type, plast_name])
    filt = df_pechat.iloc[:, [0]] == plast_name
    temp = df_pechat[filt.values.tolist()]
    if len(temp) > 1:
        temp = temp.iloc[0].values.tolist()
    else:
        temp = temp.values.tolist()
    porosity = temp[35]

    filt = df_tech.iloc[:, [4]] == w_num
    a = df_tech[filt.values.tolist()].values.tolist()
    a = a[0]
    rc = a[9] / 1000 / 2
    mu = a[44]
    B_l = a[45]
    h = a[121]
    c_t = 5e-5
    res_prm_line.extend([rc, porosity, mu, B_l, h, c_t])
    res_prms.append(res_prm_line)
    res_prm_line = []

fmts_date = ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%d.%m.%Y', '%Y.%m.%d %H:%M']
fmts_time = ['%H:%M:%S', '%H:%M']
# Считываем данные по забойному давлению
xl_P_bh = pd.ExcelFile('Проверка на реальных данных\\Отдельное (месторожд)\\day_new (добвлен еще один столбец давления).xlsx')
# xl_P_bh = pd.read_csv('Проверка на реальных данных\\Отдельное (месторожд)\\otdelnoe\\day_new (добвлен еще один столбец давления).csv',
#                       sep=';', encoding='cp1251', low_memory=False)
# df1 = xl_P_bh.parse('day')
df1 = xl_P_bh.parse('day_new')
# df1 = xl_P_bh

# создаем новый excel-файл для сохранения статистики по расчитываемым скважинам
wb = openpyxl.Workbook()

# добавляем новый лист
wb.create_sheet(title='Абсолютная ошибка', index=0)
wb.create_sheet(title='Относительная ошибка', index=1)
wb.create_sheet(title='Накопл. разность добычи-абс.', index=2)
wb.create_sheet(title='Накопл. разность добычи-отн.', index=3)
wb.create_sheet(title='Q_liq - реальный', index=4)
wb.create_sheet(title='Q_liq - модель', index=5)

# получаем лист, с которым будем работать
sheet_abs = wb['Абсолютная ошибка']
sheet_rel = wb['Относительная ошибка']
sheet_cumul = wb['Накопл. разность добычи-абс.']
sheet_cumul_rel = wb['Накопл. разность добычи-отн.']
sheet_Q_liq_act = wb['Q_liq - реальный']
sheet_Q_liq_mod = wb['Q_liq - модель']


"""-------------------------------НАЧАЛО НАСТРАИВАЕМЫХ ПАРАМЕТРОВ------------------------------------------------"""
# Индексы скважин, которые будут рассчитаны
srez_l = 0
srez_r = len(res_prms)
# srez_r = 2

comparison_start = datetime(2020, 2, 1) # Дата начала прогноза (включительно)
considered_period = 165  # Длина периода адаптации, сутки
comparison_period = 90  # Длина периода прогнозирования, сутки
fill_in_gaps = True # Заполнить пропуски дебита жидкости предыдущими значениями ПРИ ПРОГНОЗЕ
delete_common_dates = True  # Удалить общие даты для P_bh и Q_liq - иначе порой получаются выбросы в значениях модели
equal_time = True  # Заменить всю предысторию работы скважины на эквив. время ее работы

# Коэфициенты разреживания входных данных
P_bh_simplification_coeff = 12  # Данные по забойному давлению -  в период адаптации - усредняются каждые n значений, где n равен этому параметру.
Q_liq_simplification_coeff = 9  # Данные по дебиту жидкости -  в период адаптации - добавляется каждое n-ое значение, где n равен этому параметру. Остальное в рассмотрение не берется.
coef = 2    # Данные по забойному давлению -  в период прогнозирования - берется каждое n-ое значение, где n равен этому параметру.
"""-------------------------------КОНЕЦ НАСТРАИВАЕМЫХ ПАРАМЕТРОВ------------------------------------------------"""

comparison_start -= timedelta(days=1)
comparison_period += 1
considered_period += comparison_period

if equal_time:
    global df2, xl_cumul_Q
    xl_cumul_Q = pd.ExcelFile('Проверка на реальных данных\\Отдельное (месторожд)\\month.xlsx')
    df2 = xl_cumul_Q.parse('Main')
    df2 = df2.loc[1:]
    df2 = df2.drop([2, 3])
    df2.columns = df2.iloc[0, :]
    df2 = df2.drop([1])

for row in range(2, len(res_prms[srez_l:srez_r]) + 2):
    well_num = str(res_prms[srez_l + row - 2][0])
    cell_abs = sheet_abs.cell(row=row, column=1)
    cell_abs.value = well_num
    cell_rel = sheet_rel.cell(row=row, column=1)
    cell_rel.value = well_num
    cell_cumul = sheet_cumul.cell(row=row, column=1)
    cell_cumul.value = well_num
    cell_cumul_rel = sheet_cumul_rel.cell(row=row, column=1)
    cell_cumul_rel.value = well_num
    cell_Q_liq_act = sheet_Q_liq_act.cell(row=row, column=1)
    cell_Q_liq_act.value = well_num
    cell_Q_liq_mod = sheet_Q_liq_mod.cell(row=row, column=1)
    cell_Q_liq_mod.value = well_num

for col in range(2, 92):
    cell_abs = sheet_abs.cell(row=1, column=col)
    cell_abs.value = col - 1
    cell_rel = sheet_rel.cell(row=1, column=col)
    cell_rel.value = col - 1
    cell_cumul = sheet_cumul.cell(row=1, column=col)
    cell_cumul.value = col - 1
    cell_cumul_rel = sheet_cumul_rel.cell(row=1, column=col)
    cell_cumul_rel.value = col - 1
    cell_Q_liq_act = sheet_Q_liq_act.cell(row=1, column=col)
    cell_Q_liq_act.value = col - 1
    cell_Q_liq_mod = sheet_Q_liq_mod.cell(row=1, column=col)
    cell_Q_liq_mod.value = col - 1

"""-------------------------------НАЧАЛО - ОБРАБОТКА ДАННЫХ------------------------------------------------"""
for index_, elem in enumerate(res_prms[srez_l:srez_r]):
    try:
        well_num = str(elem[0])
        res_prm = elem

        well_163 = [str(elem) == well_num for elem in df1.loc[:, 'Скв']]
        well_163_date = df1.loc[:, ['Дата']]

        # well_163_P_bh_3 = df1.loc[:, ['Давление забойное от Pпр']]
        well_163_P_bh_3 = df1.loc[:, ['Давление забойное от Hд']]
        filt_P_bh = [not isnan(elem) for elem in well_163_P_bh_3[well_163].values]
        P_bh = well_163_P_bh_3[well_163][filt_P_bh].values.tolist()
        P_bh = [elem[0] for elem in P_bh]
        date_P_bh_3 = well_163_date[well_163][filt_P_bh].values
        datetimes_P_bh = [pd.to_datetime(str(datetime64[0])) for datetime64 in date_P_bh_3]

        well_163_Q_liq = df1.loc[:, ['Дебит жидкости среднесуточный']]
        a = well_163_Q_liq[well_163].values
        filt_Q_liq = [not isnan(elem) for elem in well_163_Q_liq[well_163].values]
        Q_liq = well_163_Q_liq[well_163][filt_Q_liq].values
        Q_liq = np.array([elem[0] for elem in Q_liq])
        Q_liq = np.array([float(elem) for elem in Q_liq])
        date_Q_liq_ = well_163_date[well_163][filt_Q_liq].values
        datetimes_Q_liq = [pd.to_datetime(str(datetime64[0])) for datetime64 in date_Q_liq_]

        if equal_time:
            well_163 = [str("".join(re.findall(r'\d+', elem))) == well_num for elem in df2.loc[:, 'Скважина']]
            well_163_date = df2.loc[:, ['Дата']]

            well_163_cumul_Q_ = df2.loc[:, ['Добыча жидкости']]
            filt_cumul_Q_  = [not isnan(elem) for elem in well_163_cumul_Q_[well_163].values]
            cumul_Q_ = well_163_cumul_Q_[well_163][filt_cumul_Q_].values
            cumul_Q_ = np.array([elem[0] for elem in cumul_Q_])
            cumul_Q_ = np.array([float(elem) for elem in cumul_Q_])
            date_cumul_Q_ = well_163_date[well_163][filt_cumul_Q_].values
            datetimes_cumul_Q_ = [pd.to_datetime(str(datetime64[0]), format='%m.%Y') for datetime64 in date_cumul_Q_]
    except:
        continue

    result_directory = os.path.join(
        f'RESULT\\Отдельное (месторождение)\\' + str(
            datetime.now().strftime('%Y-%m-%d-%H.%M.%S')) + ' well num ' + str(well_num))
    if not os.path.exists(result_directory):
        os.makedirs(result_directory)


    k = len(datetimes_P_bh) - 1
    while datetimes_P_bh[k] >= comparison_start + timedelta(days=comparison_period):
        k -= 1
    datetimes_P_bh = datetimes_P_bh[:k + 1]
    datetimes_P_bh.append(comparison_start + timedelta(days=comparison_period))
    P_bh = P_bh[:k + 1]
    P_bh.append(P_bh[-1])

    k = len(datetimes_Q_liq) - 1
    while datetimes_Q_liq[k] > comparison_start + timedelta(days=comparison_period):
        k -= 1
    datetimes_Q_liq = datetimes_Q_liq[:k + 1]
    Q_liq = Q_liq[:k + 1]

    delta_dates = []
    bh_pressures = []
    time_points = []
    comparison_points = []
    comparison_delta_dates = []
    start_P_bh = start_Q_liq = None
    end_P_bh = len(datetimes_P_bh)

    for i in reversed(range(len(datetimes_P_bh))):
        if datetimes_P_bh[-1] - datetimes_P_bh[i] < timedelta(days=comparison_period):
            end_P_bh = i
            continue
        if datetimes_P_bh[-1] - datetimes_P_bh[i] > timedelta(days=considered_period):
            break

    left = 0
    right = P_bh_simplification_coeff
    avg_P_bh = []

    while True:
        if right >= end_P_bh:
            avg_P_bh.append(np.mean(P_bh[left:end_P_bh]))
            break
        avg_P_bh.append(np.mean(P_bh[left:right]))
        left += P_bh_simplification_coeff
        right += P_bh_simplification_coeff
    tempo = P_bh[end_P_bh::coef]
    bh_pressures = avg_P_bh + tempo

    datetimes_P_bh = datetimes_P_bh[:end_P_bh:P_bh_simplification_coeff] + datetimes_P_bh[end_P_bh::coef]
    P_bh = P_bh[:end_P_bh:P_bh_simplification_coeff] + P_bh[end_P_bh::coef]
    for i in reversed(range(len(datetimes_P_bh))):
        if datetimes_P_bh[-1] - datetimes_P_bh[i] < timedelta(days=comparison_period):
            end_P_bh = i
            try:
                comparison_delta_dates.append((datetimes_P_bh[i + 1] - datetimes_P_bh[i]).total_seconds() / 3600)
            except:
                comparison_delta_dates.append(24)
            continue
        if datetimes_P_bh[-1] - datetimes_P_bh[i] > timedelta(days=considered_period):
            break
        try:
            delta_dates.append((datetimes_P_bh[i + 1] - datetimes_P_bh[i]).total_seconds() / 3600)
        except:
            delta_dates.append(24)
        start_P_bh = i
    delta_dates = list(reversed(delta_dates))
    comparison_delta_dates = list(reversed(comparison_delta_dates))
    bh_pressures = bh_pressures[start_P_bh:]

    if start_P_bh == None:
        continue

    Q_liq = list(Q_liq)
    for i in reversed(range(len(datetimes_Q_liq))):
        if datetimes_P_bh[-1] - datetimes_Q_liq[i] < timedelta(days=0):
            datetimes_Q_liq.pop(i)
            Q_liq.pop(i)

    if delete_common_dates:
        for elem in datetimes_Q_liq.copy():
            if elem in datetimes_P_bh[start_P_bh:end_P_bh]:
                ind = datetimes_Q_liq.index(elem)
                datetimes_Q_liq.pop(ind)
                Q_liq.pop(ind)

    end_Q_liq = len(datetimes_Q_liq)
    for i in reversed(range(len(datetimes_Q_liq))):
        if datetimes_P_bh[-1] - datetimes_Q_liq[i] >= timedelta(days=considered_period):
            break
        if datetimes_P_bh[-1] - datetimes_Q_liq[i] < timedelta(days=comparison_period):
            # if datetimes_Q_liq[i] > datetime.date(2020, 2, 1):
            end_Q_liq = i
            if datetimes_P_bh[-1] - datetimes_Q_liq[i] >= timedelta(days=0):
                continue
        start_Q_liq = i

    if equal_time:
        cumul_Q = np.sum(
            [cumul_Q_[i] for i in range(len(cumul_Q_)) if datetimes_cumul_Q_[i] < datetimes_P_bh[start_P_bh]])
        if cumul_Q != 0:
            equal_t = cumul_Q/Q_liq[start_Q_liq]    # эквивал. время работы скважины, сутки
            eq_P_bh = P_bh[start_P_bh]
            eq_Q_liq = Q_liq[start_Q_liq]
            delta_dates.insert(0, equal_t * 24)
            bh_pressures.insert(0, eq_P_bh)

    datetimes_Q_liq = datetimes_Q_liq[:end_Q_liq:Q_liq_simplification_coeff] + datetimes_Q_liq[end_Q_liq:]
    Q_liq = Q_liq[:end_Q_liq:Q_liq_simplification_coeff] + Q_liq[end_Q_liq:]

    for i in reversed(range(len(datetimes_Q_liq))):
        if datetimes_P_bh[-1] - datetimes_Q_liq[i] >= timedelta(days=considered_period) \
                or datetimes_Q_liq[i] - datetimes_P_bh[start_P_bh] < timedelta(days=0):
            break
        if datetimes_P_bh[-1] - timedelta(days=comparison_period) < datetimes_Q_liq[i] <= datetimes_P_bh[-1]:
            end_Q_liq = i
            if datetimes_P_bh[-1] - datetimes_Q_liq[i] >= timedelta(days=0):
                continue
        start_Q_liq = i

    if fill_in_gaps:    # заполнение пустых дебитов жидкости во время прогноза
        comp_indexes = [(date - comparison_start).days for date in datetimes_Q_liq[end_Q_liq:]]
        lost_indexes = [i - 1 for i in range(1, comparison_period) if i not in comp_indexes]
        new_datetimes_Q_liq = datetimes_Q_liq[end_Q_liq:].copy()
        new_Q_liq = Q_liq[end_Q_liq:].copy()
        for elem in lost_indexes:
            datetimes_Q_liq.insert(end_Q_liq + elem, comparison_start + timedelta(days=elem + 1))
            Q_liq.insert(end_Q_liq + elem, Q_liq[end_Q_liq + elem - 1])

    for i in reversed(range(len(datetimes_Q_liq))):
        if datetimes_P_bh[-1] - datetimes_Q_liq[i] >= timedelta(days=considered_period) \
                or datetimes_Q_liq[i] - datetimes_P_bh[start_P_bh] < timedelta(days=0):
            break
        if datetimes_P_bh[-1] - timedelta(days=comparison_period) < datetimes_Q_liq[i] <= datetimes_P_bh[-1]:
            end_Q_liq = i
            if datetimes_P_bh[-1] - datetimes_Q_liq[i] >= timedelta(days=0):
                if equal_time:
                    comparison_points.append(
                        (datetimes_Q_liq[i] - (datetimes_P_bh[start_P_bh] - timedelta(days=equal_t))).total_seconds() / 3600)
                else:
                    comparison_points.append(
                        (datetimes_Q_liq[i] - datetimes_P_bh[start_P_bh]).total_seconds() / 3600)
            continue
        if equal_time:
            time_points.append((datetimes_Q_liq[i] - (datetimes_P_bh[start_P_bh] - timedelta(days=equal_t))).total_seconds() / 3600)
        else:
            time_points.append((datetimes_Q_liq[i] - datetimes_P_bh[start_P_bh]).total_seconds() / 3600)
        start_Q_liq = i

    time_points = list(reversed(time_points))
    comparison_points = list(reversed(comparison_points))
    # try:
    #     time_points.append(comparison_points[0])
    # except:
    #     pass
    considered_length = len(time_points)

    """-------------------------------КОНЕЦ - ОБРАБОТКА ДАННЫХ------------------------------------------------"""


    WM = {'ВЕРТ': 0, 'ГОР': 2, 'ннсгрп': 2, 'mfrac': 3}
    BoundType = 0
    Q = []
    Q_comparison = []
    best = 1
    x_ = 0
    num_calls = 0
    Q_arr = np.array([])
    Q_arr_comp = np.array([])

    # k=x[0], skin=x[1], w=x[2], l=x[3], P_initial=x[4], xf=x[5], Fc=x[6]
    if WM[res_prm[1]] == 2:
        bounds = [(0.1, 104), (0.0001, 10), (200, 26500), (200, 26500), (max(bh_pressures), max(bh_pressures) * 3),
                  (0.1, 100), (100, 10000)]
    elif WM[res_prm[1]] == 0:
        bounds = [(0.1, 104), (0.0001, 10), (200, 26500), (200, 26500), (max(bh_pressures), max(bh_pressures) * 3)]

    weight = None

    results_x = []
    results_q = []
    results_fun = []
    results_num_calls = []
    results_exec_time = []
    results_alg_name = []
    results_P_initial = []
    start_time = time.time()

    """-------------------------------НАЧАЛО ОПТИМИЗАЦИИ------------------------------------------------"""
    # ----------GPSE-----------
    params = {'m': 3, 'step_size_coeff': 10}
    options = {'maxiter': 60, 'final_sten_size': 0.000001}
    for i in range(1):
        if WM[res_prm[1]] == 2:
            bounds = [(0.1, 104), (0.0001, 10), (200, 26500), (200, 26500),
                      (max(bh_pressures), max(bh_pressures) * 3), (0.1, 100), (100, 10000)]
        elif WM[res_prm[1]] == 0:
            bounds = [(0.1, 104), (0.0001, 10), (200, 26500), (200, 26500),
                      (max(bh_pressures), max(bh_pressures) * 3)]

        q_ = 1
        start_weight = 1 / q_ ** len(time_points)
        exp_weight = []
        for u in range(1, len(time_points) + 1):
            exp_weight.append(start_weight * q_ ** u)
        weight = np.array(exp_weight)
        # weight = [1 for i in range(len(time_points))]
        num_calls = 0
        t3 = time.time()
        x_init = None
        x_inits = []
        lcube = lhsmdu.sample(len(bounds), 100, randomSeed=random.randint(1, 10000)).transpose()
        for elem in lcube:
            a = elem.tolist()[0]
            x_sample = []
            for i in range(len(bounds)):
                x_sample.append(bounds[i][0] + (bounds[i][1] - bounds[i][0]) * a[i])
            x_inits.append(x_sample)
        t = 100000
        for elem in x_inits:
            sample_loss = loss_func(elem)
            if sample_loss < t:
                t = sample_loss
                x_init = elem

        res = gpse.gpse(fun=loss_func_maximize, x0=x_init, bounds=bounds, options=options, params=params)
        x = res[0]
        loss_func(x)
        save_plot('Fixed_p_init', 1)

        q_ = 10
        start_weight = 1 / q_ ** len(time_points)
        exp_weight = []
        for i in range(1, len(time_points) + 1):
            exp_weight.append(start_weight * q_ ** i)
        weight = np.array(exp_weight)
        bounds = [(max(bh_pressures), max(bh_pressures) * 4)]
        x_init = None
        x_inits = []
        lcube = lhsmdu.sample(1, 5, randomSeed=random.randint(1, 10000)).transpose()
        for elem in lcube:
            a = elem.tolist()[0]
            a = a[0]
            x_sample = []
            x_sample.append(bounds[0][0] + (bounds[0][1] - bounds[0][0]) * a)
            x_inits.append(x_sample)
        t = 100000
        for elem in x_inits:
            sample_loss = loss_func_maximize_P_init_only(elem)
            if sample_loss < t:
                t = sample_loss
                x_init = elem
        params = {'m': 3, 'step_size_coeff': 10}
        options = {'maxiter': 50, 'final_sten_size': 0.00001}
        res = gpse.gpse(fun=loss_func_maximize_P_init_only, x0=x_init, bounds=bounds, options=options, params=params)
        p_init = res[0]
        x[4] = p_init[0]
        results_q.append(q_)
        results_x.append(x)
        results_fun.append(res[1])
        results_num_calls.append(len(res[2]))
        results_exec_time.append(time.time() - t3)
        results_alg_name.append('-GPSE- ')
        loss_func_maximize_P_init_only(res[0])
        save_plot('Adaptation', q_)

        delta_dates_ = delta_dates + comparison_delta_dates
        pw = bh_pressures
        pw.append(pw[-1])
        time_moments = [np.sum(delta_dates_[:i + 1]) for i in range(len(delta_dates_))]
        time_moments.insert(0, 0)
        delta_p = [bh_pressures[0] - bh_pressures[i] for i in range(1, len(bh_pressures))]
        delta_p.append(delta_p[-1])
        slope = [-(pw[i + 1] - pw[i]) / delta_dates_[i] for i in range(len(delta_dates_))]
        free_term = [x[4] - (pw[i] * time_moments[i + 1] - pw[i + 1] * time_moments[i]) / delta_dates_[i] for i in
                     range(len(delta_dates_))]

        input_list = [(delta_dates_[i] / 24, delta_p[i], slope[i], free_term[i]) for i in range(len(delta_dates_))]

        Q = np.array([])
        Q_comparison = np.array([])
        for t in time_points:
            try:
                q = Pwf_Ql_Qt(t=t, Regimes=input_list, k=x[0], regimes_flag=2,
                              skin=x[1], h=res_prm[7], ct=res_prm[8], mu=res_prm[5], bl=res_prm[6],
                              rw=res_prm[3], phi=res_prm[4], Wellmodel=WM[res_prm[1]],
                              ReservoirShape=0, BoundaryType=BoundType, xf=x[5],
                              Fc=x[6],
                              w=x[2], l=x[3], wf=0.5, lf=0.5)
            except:
                q = Pwf_Ql_Qt(t=t, Regimes=input_list, k=x[0], regimes_flag=2,
                              skin=x[1], h=res_prm[7], ct=res_prm[8], mu=res_prm[5], bl=res_prm[6],
                              rw=res_prm[3], phi=res_prm[4], Wellmodel=WM[res_prm[1]],
                              ReservoirShape=0, BoundaryType=BoundType,
                              w=x[2], l=x[3], wf=0.5, lf=0.5)
            Q = np.append(Q, q)
        for t in comparison_points:
            try:
                q = Pwf_Ql_Qt(t=t, Regimes=input_list, k=x[0], regimes_flag=2,
                              skin=x[1], h=res_prm[7], ct=res_prm[8], mu=res_prm[5], bl=res_prm[6],
                              rw=res_prm[3], phi=res_prm[4], Wellmodel=WM[res_prm[1]],
                              ReservoirShape=0, BoundaryType=BoundType, xf=x[5],
                              Fc=x[6],
                              w=x[2], l=x[3], wf=0.5, lf=0.5)
            except:
                q = Pwf_Ql_Qt(t=t, Regimes=input_list, k=x[0], regimes_flag=2,
                              skin=x[1], h=res_prm[7], ct=res_prm[8], mu=res_prm[5], bl=res_prm[6],
                              rw=res_prm[3], phi=res_prm[4], Wellmodel=WM[res_prm[1]],
                              ReservoirShape=0, BoundaryType=BoundType,
                              w=x[2], l=x[3], wf=0.5, lf=0.5)
            Q_comparison = np.append(Q_comparison, q)
        save_plot('Prediction-', q_, prediction=True)
        save_hist('Prediction-', q_)
        save_cumul('Prediction-', q_)
        dev = (Q - Q_liq[start_Q_liq:end_Q_liq]) * (Q - Q_liq[start_Q_liq:end_Q_liq])
        erro = np.sqrt(np.sum((dev * weight)) / considered_length)
        if len(Q_comparison) > 0:
            write_excel(index_ + 2)

    # ----------Diff. Evolution-----------
    for i in range(500, 500):
        num_calls = 0
        t1 = time.time()
        result = differential_evolution(loss_func, bounds, maxiter=8, popsize=2 * (len(bounds) + 1), workers=1,
                                        disp=True, callback=check_convergence(), polish=False)
        results_q.append(q_)
        results_x.append(result.x)
        results_fun.append(result.fun)
        results_num_calls.append(num_calls)
        results_exec_time.append(time.time() - t1)
        results_alg_name.append('-Diff.Ev.- ')
        results_P_initial.append(result.x[4])
        print('      P_initial =  ' + str(result.x[4]))
        loss_func(result.x)
        save_plot('-Diff.Ev.', q_)

        x = result.x
        delta_dates_ = delta_dates + comparison_delta_dates
        bh_pressures_ = bh_pressures
        input_list = [(delta_dates_[i] / 24, x[4] - bh_pressures_[i]) for i in range(len(delta_dates_))]
        Q = np.array([])
        Q_comparison = np.array([])
        for t in time_points:
            try:
                q = Pwf_Ql_Qt(t=t, Regimes=input_list, k=x[0], regimes_flag=2,
                              skin=x[1], h=res_prm[7], ct=res_prm[8], mu=res_prm[5], bl=res_prm[6],
                              rw=res_prm[3], phi=res_prm[4], Wellmodel=WM[res_prm[1]],
                              ReservoirShape=0, BoundaryType=BoundType, xf=x[5],
                              Fc=x[6],
                              w=x[2], l=x[3], wf=0.5, lf=0.5)
            except:
                q = Pwf_Ql_Qt(t=t, Regimes=input_list, k=x[0], regimes_flag=2,
                              skin=x[1], h=res_prm[7], ct=res_prm[8], mu=res_prm[5], bl=res_prm[6],
                              rw=res_prm[3], phi=res_prm[4], Wellmodel=WM[res_prm[1]],
                              ReservoirShape=0, BoundaryType=BoundType,
                              w=x[2], l=x[3], wf=0.5, lf=0.5)
            Q = np.append(Q, q)
        for t in comparison_points:
            try:
                q = Pwf_Ql_Qt(t=t, Regimes=input_list, k=x[0], regimes_flag=2,
                              skin=x[1], h=res_prm[7], ct=res_prm[8], mu=res_prm[5], bl=res_prm[6],
                              rw=res_prm[3], phi=res_prm[4], Wellmodel=WM[res_prm[1]],
                              ReservoirShape=0, BoundaryType=BoundType, xf=x[5],
                              Fc=x[6],
                              w=x[2], l=x[3], wf=0.5, lf=0.5)
            except:
                q = Pwf_Ql_Qt(t=t, Regimes=input_list, k=x[0], regimes_flag=2,
                              skin=x[1], h=res_prm[7], ct=res_prm[8], mu=res_prm[5], bl=res_prm[6],
                              rw=res_prm[3], phi=res_prm[4], Wellmodel=WM[res_prm[1]],
                              ReservoirShape=0, BoundaryType=BoundType,
                              w=x[2], l=x[3], wf=0.5, lf=0.5)
            Q_comparison = np.append(Q_comparison, q)
        save_plot('-Diff.Ev.-prediction-', q_, prediction=True)
        save_hist('-Diff.Ev.-prediction-', q_)
        save_cumul('-Diff.Ev.-prediction-', q_)
        Q_second = np.copy(Q)
        Q_second_comp = np.copy(Q_comparison)
    print('exec time : ' + str(time.time() - start_time))

    """-------------------------------КОНЕЦ ОПТИМИЗАЦИИ------------------------------------------------"""

    if len(Q_comparison) > 0:
        write_excel(index_ + 2)

    """-------------------------------Запись лога в текстовый файл по каждой скважине------------------------------------------------"""
    if True:
        f = open(os.path.join(result_directory, str(well_num) + ' - log.txt'), 'w')
        for i in range(len(results_fun)):
            f.write(str(results_q[i]) + str(results_alg_name[i]) + '\n')
            f.write('\tx_optimum = ' + str(results_x[i]) + '\n')
            f.write('\tloss function value = ' + str(abs(results_fun[i])) + '\n')
            f.write('\titerations = ' + str(results_num_calls[i]) + '\n')
            f.write('\tWell num: ' + str(well_num) + '\n')
            f.write('\tWell model: ' + str(res_prm[1]) + '\n')
            f.write('\texec_time, sec = ' + str(results_exec_time[i]) + '\n')
            f.write('\n\n')
        f.write('final error: ' + str(erro) + '\n')
        f.write('\n\n')
        f.close()

wb.save(os.path.join(result_directory, "..", 'Ошибка адаптации.xlsx'))

